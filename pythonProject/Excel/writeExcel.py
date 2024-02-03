import openpyxl
from pathlib import Path

# create an Excel file
excelFile = openpyxl.Workbook()
print(excelFile.sheetnames)

# change sheet name
excelSheet = excelFile.active
excelSheet.title = 'firstSheet'

# create sheet
excelFile.create_sheet()  # this creates new sheets in the Excel file
excelFile.create_sheet()

# this creates a new sheet but in a specific index in the Excel sheet
excelFile.create_sheet(index=1, title='secondSheet')
# create a sheet and place it after the secondSheet
excelFile.create_sheet(index=2, title='thirdSheet')

# delete sheet
del excelFile['thirdSheet']  # this deletes the sheet from the Excel sheets

# write to an Excel sheet
sheet = excelFile['secondSheet']
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)

# exercise: write four names in the third column (this how I did it)
sheet = excelFile['firstSheet']
sheet['C1'] = 'Hamada'
sheet['C2'] = 'Amira'
sheet['C3'] = 'Yassin'
sheet['C4'] = 'Frida'
for i in range(1, 5):
    print(i, sheet.cell(row=i, column=3).value)

# this is how he did the same exercise
names = ['Hamada', 'Amira', 'Yassin', 'Frida']
firstSheet = excelFile['secondSheet']  # I just printed them in another place to know the difference
for i in range(1, len(names)+1):  # +1 to get the last index of the file
    firstSheet.cell(row=i, column=3).value = names[i-1]

# save Excel file
excelFile.save(filename=Path.home() / Path("OneDrive", "Escritorio") / 'newExcel.xlsx')
