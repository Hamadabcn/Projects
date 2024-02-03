import openpyxl
from pathlib import Path

# to open an Excel file
excelFile = openpyxl.load_workbook(Path.home() / Path("OneDrive", "Escritorio", "example.xlsx"))
print(excelFile.sheetnames)  # this will print the Excel sheet names

# this to read a specific sheet
Sheet1 = excelFile['Sheet1']
print(Sheet1.title)

activeSheet = excelFile.active
print(activeSheet.title)

print(Sheet1['A1'].value)  # this prints the value in a specific cell
print(Sheet1['B1'].value)
print(Sheet1['C1'].value)

print(Sheet1['C1'].row)  # this prints the number of the row the cell is in
print(Sheet1['C1'].column)  # this prints the number of the column the cell is in

print(Sheet1['C1'].coordinate)  # this prints the name of the cell you are working with

print(Sheet1.cell(row=1, column=2).value)  # this prints the value of a specific cell
print('-' * 50)
# what if you want to print all the names in the Excel sheet
for i in range(1, 7):
    print(i, Sheet1.cell(row=i, column=1).value)
print('-' * 50)

# what if you want to print all the salaries in the Excel sheet
for i in range(1, 7):
    print(i, Sheet1.cell(row=i, column=2).value)
print('-' * 50)

# what if you want to print all the names, salaries, and the total amount of salaries the company pays
total = 0
for i in range(1, Sheet1.max_row+1):  # +1 so it gets the last index of the sheet
    print(i, Sheet1.cell(row=i, column=1).value, Sheet1.cell(row=i, column=2).value)
    total += Sheet1.cell(row=i, column=2).value
print(f'The total salaries of the employees are {total} €')
print('-' * 50)

print(Sheet1.max_row)  # this is to know the total number of rows in the Excel sheet
print(Sheet1.max_column)  # this is to know the total number of columns it the Excel sheet

